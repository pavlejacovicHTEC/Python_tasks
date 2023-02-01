import argparse
import os
import datetime
import pathlib
import sys
from openpyxl import load_workbook, Workbook
import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.drawing.fill import PatternFillProperties
from openpyxl.chart.marker import DataPoint


parser = argparse.ArgumentParser()
parser.add_argument('-s', '--start_date', type=str, help="Start date for history charts")
parser.add_argument('-e', '--end_date', type=str, help="End date to history charts")
parser.add_argument('-n', '--nightly', action='store_true', help="Add last 10 results chart to nightly report")
parser.add_argument('-r', '--report', type=str, help="Path to test reports")
parser.add_argument('-t', '--test', type=str, required=True, help="Path to test results")
args = parser.parse_args()
date_time_format = '%Y-%m-%d'

#date checks
if (args.start_date and not args.end_date) or (not args.start_date and args.end_date):
    sys.exit("You must provide start and end date together")
if args.start_date and args.end_date:
    result = True
    try:
        result = bool(datetime.datetime.strptime(args.start_date, date_time_format))
        result = bool(datetime.datetime.strptime(args.end_date, date_time_format))
    except ValueError:
        sys.exit('Start and end date must be in valid format ---> year-month-day')


#this checks if you provided combination of dates and nightly argument
if args.nightly and args.start_date:
    sys.exit("You can not combine dats and night arguments")
#this checks if reports path is provided when nightly argument is provided, that is mandatory
if (args.nightly and not args.report) or (not args.nightly and args.report):
    sys.exit("You must specify argument <excel path> and argument <night> together")


#this function return commit_id if there is one
def get_commit(path):
    current_dir = os.getcwd()
    os.chdir(path)

    with open('commit.txt') as commit_file:
        commit_id = str(commit_file.readlines(1))[9:15]
    commit_file.close()

    os.chdir(current_dir)
    return commit_id


#this function return date_time and commit_id as one string
def get_date_time_commit_id(folder):
    return str(folder.split('-')[-5]) + '-' + \
           str(folder.split('-')[-4]) + '-' + \
           str(folder.split('-')[-3]) + '  \n' + \
           get_commit(folder)


def get_last_file_in_report():
    new_path = sorted(pathlib.Path(args.report).glob('*xlsx'), key=os.path.getmtime)[-1]
    return str(new_path)


#this function returns n-th modified folder path that contains perf in it's name
def get_last_nth(n, path):
    new_path = sorted(pathlib.Path(path).glob('perf*'), key=os.path.getmtime)[-n]
    return new_path


#delets every model sheet
def purge_model_sheets(wb):
    sheets_list = wb.sheetnames

    for element in range(len(sheets_list)):
        if element != 0:
            del wb[sheets_list[element]]

    wb.save(get_last_file_in_report())


#this function returns folders that are modifed between two dates
def get_folders_between_dates():
    targeted_folders = []

    for member in os.listdir(args.test):
        element = os.path.join(args.test, member)
        sub_elements = member.split('-')

        if os.path.isdir(element) and "perf" == sub_elements[0]:
            date_string = sub_elements[1] + "-" + \
                          sub_elements[2] + "-" + \
                          sub_elements[3]
            target_date = datetime.datetime.strptime(date_string, date_time_format)
            start_date = datetime.datetime.strptime(args.start_date, date_time_format)
            end_date = datetime.datetime.strptime(args.end_date, date_time_format)
            if start_date <= target_date <= end_date and os.path.isfile(element + os.sep + "results.csv"):
                targeted_folders.append(element)

    return targeted_folders


#returns last x modified folders
def get_last_x_folders(x):
    last_x_folders = []
    counter = 0
    temp = 1

    while temp <= x:
        counter += 1
        path = str(get_last_nth(counter, args.test))
        if os.path.isfile(path+os.path.sep+"results.csv"):
            last_x_folders.append(path)
            temp += 1

    return last_x_folders


#this function returns all the model names from results.csv file in specified folders
def get_all_model_names(folders):
    model_names_list = []
    counter = 0

    while counter < len(folders):
        model_names_list_temp = pd.read_csv(folders[counter] + os.sep + "results.csv", header=None,usecols=[0, 1]).values.tolist()
        for element in model_names_list_temp:
            element = str(element).replace('[', '').replace(']', '').replace(' ', '').replace('\'', '')
            if element not in model_names_list:
                model_names_list.append(element)
        counter += 1

    return model_names_list


#this function calls the functions that returns last 10 folders and then it colects all results from each model into a dictionary
def get_results_for_each_model_for_last_x_folders(x):
    folder_list = get_last_x_folders(10)
    model_results_dictionary = {}

    for model_name in get_all_model_names(get_last_x_folders(x)):
        model_values_list = []
        for folder in folder_list:
            with open(folder+os.path.sep+"/results.csv") as results_file:
                for line in results_file:
                    if model_name in line:
                        result_value = line.split(",")[2][:-1]
                        model_values_list.append(result_value)
            results_file.close()

        model_results_dictionary[model_name] = model_values_list

    return model_results_dictionary

#defining chart parameters
def parametrize_chart(sheet):
    chart = BarChart()
    chart.type = "col"
    chart.style = 3
    chart.title = (f'{sheet}'.replace("<Worksheet \"", "").replace(">", "").replace('"', ""))
    chart.x_axis.title = 'Date'
    chart.y_axis.title = "Rate"
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.height = 10
    chart.width = 37
    chart.legend = None
    data = Reference(sheet, min_col=2, min_row=2, max_col=2, max_row=sheet.max_row)
    cats = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)
    chart.add_data(data)
    chart.set_categories(cats)
    pt = DataPoint(idx=0)
    s1 = chart.series[0]
    pt.graphicalProperties.solidFill = "ff8400"
    s1.dPt.append(pt)
    sheet.add_chart(chart, "F2")


#adds hyperlinks to each model so it points to right model sheet
def link_main_table_to_model_sheets():
    wb = load_workbook(get_last_file_in_report())
    target_doc_basename = os.path.basename(get_last_file_in_report())
    target_sheet = wb[target_doc_basename.strip(".xlsx")]

    counter = 5
    while counter < target_sheet.max_row:
        target_cell_a = target_sheet[f"A{counter}"]
        target_cell_b = target_sheet[f"B{counter}"]
        full_model_name = str(target_cell_a.value) + "," + str(target_cell_b.value)
        link_to_model_sheet = f"{target_doc_basename}#'{full_model_name}'!A1"
        target_sheet.cell(row=counter, column=1).hyperlink = link_to_model_sheet
        counter += 1
    wb.save(get_last_file_in_report())


#adds go back button
def go_back_button(target_sheet):
    main_doc_basename = os.path.basename(get_last_file_in_report())
    main_sheet_name = main_doc_basename.strip(".xlsx")

    link_to_main_sheet = f"{main_doc_basename}#'{main_sheet_name}'!A1"

    fill_cell = PatternFill(patternType='solid', fgColor='C64747')
    target_sheet.cell(row=20, column=1, value="Go back").hyperlink = link_to_main_sheet
    target_sheet.cell(row=20, column=1).fill = fill_cell
    target_sheet.cell(row=20, column=1).font = Font(bold=True, size=15)


#this function creates excel graphs for models on seperate sheets
def create_excel_graphs_between_wanted_dates(folders):
    wb = Workbook()
    del wb['Sheet']
    max_model_names_list = get_all_model_names(get_folders_between_dates())

    #creates sheets for all the existing models
    for model_name in max_model_names_list:
        model_sheet = wb.create_sheet(model_name)
        model_sheet.append(["Date & commit_ID"] + ["Value"])

    #goes thorugh folders and takes each model and writes it's results to particular sheet
    for folder in folders:
        model_names_list = pd.read_csv(folder + os.sep + "results.csv", header=None, usecols=[0, 1]).values.tolist()
        results_list = pd.read_csv(folder + os.sep + "results.csv", header=None, usecols=[2]).values.tolist()
        date_time_commit_id_column1 = get_date_time_commit_id(folder)

        counter = 0
        for element in model_names_list:
            result_column2 = float(str(results_list[counter]).replace('[', '').replace(']', ''))
            target_sheet = wb[str(element).replace('[', '').replace(']', '').replace(' ', '').replace('\'', '')]
            target_sheet.append([date_time_commit_id_column1, result_column2])
            target_sheet.column_dimensions['A'].width = 20
            counter += 1

    for sheet in wb:
        parametrize_chart(sheet)

    wb.save('test_results_with_models_separated.xlsx')


#this function creates excel graphs for last x result
def create_excel_graphs_for_last_x_results(x):
    wb = load_workbook(get_last_file_in_report())
    last_ten_folders = get_last_x_folders(10)

    purge_model_sheets(wb)

    for key, value in get_results_for_each_model_for_last_x_folders(x).items():
        counter = 0
        model_sheet = wb.create_sheet(key)
        model_sheet.append(["Date & commit_ID"]+["Value"])
        target_sheet = wb[key]
        target_sheet.column_dimensions['A'].width = 20
        for result_column2 in value:
            date_time_commit_id_column1 = get_date_time_commit_id(last_ten_folders[counter])
            target_sheet.append([date_time_commit_id_column1, result_column2])
            counter += 1
            for row in range(2, target_sheet.max_row + 1):
                cell_value = target_sheet.cell(row=row, column=2).value
                if cell_value is None:
                    continue
                float_value = pd.to_numeric(cell_value, errors='coerce')
                target_sheet.cell(row=row, column=2, value=float_value)
            parametrize_chart(target_sheet)
        go_back_button(target_sheet)
    wb.save(get_last_file_in_report())
    link_main_table_to_model_sheets()


def main():
    if not args.nightly:
        create_excel_graphs_between_wanted_dates(get_folders_between_dates())
    else:
        create_excel_graphs_for_last_x_results(10)
    print("Report updated!")


#this is where everything is called
main()
